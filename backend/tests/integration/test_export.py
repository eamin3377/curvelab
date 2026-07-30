"""Integration tests for the export endpoints."""

from __future__ import annotations

import base64
import io
import json
import zipfile

import pytest
from fastapi.testclient import TestClient

SPRING = {
    "x": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
    "y": [2.9, 5.1, 6.8, 9.2, 10.9, 13.1, 14.8, 17.2, 18.9, 21.1, 22.8, 25.2],
}

def _make_png_b64() -> str:
    """Generate a small valid PNG in-memory for the chart-embed path."""
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (48, 32), (79, 70, 229)).save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


TINY_PNG = _make_png_b64()


def _payload(**overrides) -> dict:
    body = {
        "fit_request": {**SPRING, "model": "linear"},
        "report_meta": {
            "title": "Test Report",
            "author": "A. Student",
            "institution": "Test University",
            "course": "Numerical Methods",
        },
        "chart_png_base64": TINY_PNG,
    }
    body.update(overrides)
    return body


def test_export_csv(client: TestClient) -> None:
    """CSV export contains input, calculation and metrics sections."""
    res = client.post("/api/v1/export/csv", json=_payload())
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("text/csv")
    assert "attachment" in res.headers["content-disposition"]
    text = res.content.decode("utf-8-sig")
    assert "# Input Data" in text and "# Calculation Table" in text
    assert "# Metrics" in text and "SUM" in text


def test_export_json_roundtrip(client: TestClient) -> None:
    """JSON export is parseable and mirrors the fit contract."""
    res = client.post("/api/v1/export/json", json=_payload())
    assert res.status_code == 200
    doc = json.loads(res.content)
    assert doc["meta"]["title"] == "Test Report"
    assert doc["input"]["x"] == SPRING["x"]
    assert doc["result"]["metrics"]["r2"] > 0.999


def test_export_txt(client: TestClient) -> None:
    """TXT export is a readable report with the fitted equation."""
    res = client.post("/api/v1/export/txt", json=_payload())
    assert res.status_code == 200
    text = res.content.decode()
    assert "CURVELAB" in text
    assert "y = 0.9591" in text.replace("0.959", "0.9591")[: len(text)]
    assert "STEP-BY-STEP" in text.upper()


def test_export_xlsx(client: TestClient) -> None:
    """XLSX export opens with openpyxl and has the documented sheets."""
    res = client.post("/api/v1/export/xlsx", json=_payload())
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    assert {"Input", "Calculation", "Results", "Summary", "Graph"} <= set(wb.sheetnames)
    assert wb["Input"].max_row == 13  # header + 12 points


def test_export_pdf(client: TestClient) -> None:
    """PDF export returns a real, non-trivial PDF document."""
    res = client.post("/api/v1/export/pdf", json=_payload())
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert res.content[:5] == b"%PDF-"
    assert len(res.content) > 3000


def test_export_docx(client: TestClient) -> None:
    """DOCX export is a valid Word package with report content."""
    res = client.post("/api/v1/export/docx", json=_payload())
    assert res.status_code == 200
    assert zipfile.is_zipfile(io.BytesIO(res.content))
    from docx import Document

    doc = Document(io.BytesIO(res.content))
    full_text = "\n".join(p.text for p in doc.paragraphs)
    assert "Test Report" in full_text
    assert "Step-by-Step" in full_text


def test_export_rejects_bad_chart(client: TestClient) -> None:
    """A non-PNG chart payload is a schema-level 422."""
    bad = _payload(chart_png_base64=base64.b64encode(b"not-a-png").decode())
    res = client.post("/api/v1/export/pdf", json=bad)
    assert res.status_code == 422


def test_export_unknown_format(client: TestClient) -> None:
    """Unknown formats return the documented 400 problem."""
    res = client.post("/api/v1/export/bmp", json=_payload())
    assert res.status_code == 400
    assert res.json()["type"] == "unknown_export_format"


def test_export_filename_has_model_and_extension(client: TestClient) -> None:
    """Content-Disposition carries a sanitized, timestamped filename."""
    res = client.post("/api/v1/export/pdf", json=_payload())
    disposition = res.headers["content-disposition"]
    assert "curvelab_linear_" in disposition and disposition.endswith('.pdf"')


@pytest.mark.parametrize("model", ["linear", "polynomial", "exponential"])
def test_export_pdf_all_models(client: TestClient, model: str) -> None:
    """Every model exports a valid PDF (exponential needs positive y)."""
    body = _payload()
    body["fit_request"] = {
        "x": [0, 1, 2, 3, 4, 5],
        "y": [1.2, 2.1, 3.9, 8.2, 16.1, 31.9],
        "model": model,
        "degree": 2,
    }
    res = client.post("/api/v1/export/pdf", json=body)
    assert res.status_code == 200
    assert res.content[:5] == b"%PDF-"
