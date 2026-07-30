"""Excel export: a styled multi-sheet workbook built with openpyxl.

Sheets: Input | Calculation | Results | Summary (+ Graph when a chart PNG
is embedded). Headers use the brand indigo, panes are frozen, and the
summation row is emphasized — the workbook is meant to be presented, not
just opened.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.engines.export.base import ExportedFile, ExportPayload, exporter

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
SUM_FILL = PatternFill("solid", fgColor="EEF2FF")
HEADER_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1E1B4B")
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="475569")
BODY_FONT = Font(name="Consolas", size=10)
THIN_BOTTOM = Border(bottom=Side(style="thin", color="E2E8F0"))

MAX_SHEET_ROWS = 1000


def _style_header_row(ws, row: int, ncols: int) -> None:
    """Apply the indigo header style to a worksheet row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, widths: dict[int, float]) -> None:
    """Set column widths from a {column_index: width} mapping."""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


@exporter("xlsx")
def build_xlsx(payload: ExportPayload) -> ExportedFile:
    """Render the fit as a formatted Excel workbook and return its bytes."""
    fit = payload.fit
    wb = Workbook()

    # ── Input ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Input"
    ws.append(["i", "x", "y"])
    _style_header_row(ws, 1, 3)
    for i, (xv, yv) in enumerate(zip(payload.x, payload.y), start=1):
        ws.append([i, xv, yv])
    ws.freeze_panes = "A2"
    _autofit(ws, {1: 6, 2: 14, 3: 14})

    # ── Calculation ────────────────────────────────────────────────────
    ws = wb.create_sheet("Calculation")
    table = fit.calculation_table
    ws.append(table.columns)
    _style_header_row(ws, 1, len(table.columns))
    for row in table.rows[:MAX_SHEET_ROWS]:
        ws.append(row)
    sum_row = ws.max_row + 1
    ws.cell(row=sum_row, column=1, value="SUM")
    for j, value in enumerate(table.sums, start=1):
        cell = ws.cell(row=sum_row, column=j + 1, value=round(value, 8))
        cell.fill = SUM_FILL
        cell.font = Font(name="Consolas", size=10, bold=True, color="4338CA")
    ws.cell(row=sum_row, column=1).fill = SUM_FILL
    ws.cell(row=sum_row, column=1).font = Font(bold=True, color="4338CA")
    if table.truncated:
        ws.append([])
        ws.append([f"Note: first {MAX_SHEET_ROWS} of {table.total_rows} rows shown."])
    ws.freeze_panes = "A2"
    _autofit(ws, {i: 13 for i in range(1, len(table.columns) + 1)})

    # ── Results ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Results")
    ws["A1"] = "CurveLab — Fit Results"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Model"
    ws["B3"] = f"{fit.model.value} (degree {fit.degree})"
    ws["A4"] = "Equation"
    ws["B4"] = fit.equation.plain
    for r in ("A3", "A4"):
        ws[r].font = LABEL_FONT
    ws.append([])
    ws.append(["Coefficient", "Value"])
    _style_header_row(ws, 6, 2)
    for c in fit.coefficients:
        ws.append([c.name, round(c.value, 10)])
    ws.append([])
    start = ws.max_row + 1
    ws.cell(row=start, column=1, value="Metric")
    ws.cell(row=start, column=2, value="Value")
    _style_header_row(ws, start, 2)
    for key, value in vars(fit.metrics).items():
        ws.append([key.upper(), round(value, 10) if value is not None else "—"])
    _autofit(ws, {1: 16, 2: 28})

    # ── Summary ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    meta = payload.meta
    ws["A1"] = meta.title
    ws["A1"].font = TITLE_FONT
    rows = [
        ("Institution", meta.institution),
        ("Course", meta.course),
        ("Author", meta.author),
        ("Student ID", meta.student_id),
        ("Date", meta.date),
        ("Data points", fit.n),
        ("Equation", fit.equation.plain),
        ("R²", round(fit.metrics.r2, 6)),
    ]
    for label, value in rows:
        if value in ("", None):
            continue
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = LABEL_FONT
        ws.cell(row=ws.max_row, column=1).border = THIN_BOTTOM
        ws.cell(row=ws.max_row, column=2).border = THIN_BOTTOM
    _autofit(ws, {1: 16, 2: 44})

    # ── Graph (optional embedded chart) ────────────────────────────────
    if payload.chart_png:
        ws = wb.create_sheet("Graph")
        img = XlImage(io.BytesIO(payload.chart_png))
        img.width = min(img.width, 960)
        img.height = int(img.height * (img.width / max(img.width, 1)))
        ws.add_image(img, "B2")

    buf = io.BytesIO()
    wb.save(buf)
    return ExportedFile(
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    )
